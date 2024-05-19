import openpyxl
import pandas as pd

file_path = "excel_test/text1.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb["Sheet1"]
all_merged_cell_ranges = list(sheet.merged_cells.ranges)
print(all_merged_cell_ranges)
for merged_cell_range in all_merged_cell_ranges:
    # 遍历所有的合并区间，得到合并区间的值
    merger_cell = merged_cell_range.start_cell
    print("合并区域的值：", merger_cell, merger_cell.value)
    # sheet.unmerge_cells(merged_cell_range)
    # for row_index, col_index in merged_cell_range.cells:
    #     cell = sheet.cell(row=row_index, column=col_index)
    #     cell.value = merger_cell.value
    # 将合并区间取消合并，并将对应的单元格内填入对应的值
    min_col, min_row, max_col, max_row = merged_cell_range.bounds
    print(min_col, min_row, max_col, max_row)
    range_string = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    print("range_string = ", range_string)
    # 取消合并单元格
    sheet.unmerge_cells(range_string)
    # 取消合并后单元格中填入对应的值
    for i in range(min_row, max_row + 1):
        for j in range(min_col, max_col + 1):
            sheet.cell(row=i, column=j).value = merger_cell.value

temp_file_path = file_path.replace(".xlsx", "unmerged.xlsx")
wb.save(temp_file_path)
# 使用pandas读取保存后的excel信息，df是一个DataFrame类型的对象
df = pd.read_excel(temp_file_path, sheet_name="Sheet1")
print(type(df), df)
print(df.to_dict(orient="records"))
# 如何判断一个字符串是否在一个DataFrame对象中被包含
